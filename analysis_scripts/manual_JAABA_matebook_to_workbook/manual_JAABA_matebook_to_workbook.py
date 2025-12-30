# Last edited by Paulami Dey on 28/11/2024

import os
import csv
import glob
import openpyxl

# Path to the main folder
main_folder = input("What is the main folder's path?")

# Paths to subfolders
first_csv_folder = os.path.join(main_folder, "manual_output_files")
second_csv_folder = os.path.join(main_folder, "JAABA_output")
third_csv_folder = os.path.join(main_folder, "matebook_output")

# Find the workbook file that ends with "Workbook.xlsx"
workbook_files = glob.glob(os.path.join(main_folder, "*Workbook.xlsx"))

# Check if any workbook file is found
if not workbook_files:
    raise FileNotFoundError("No workbook file ending with 'Workbook.xlsx' found in the specified directory.")

# Use the first matched workbook file (you can modify this if needed)
workbook_path = workbook_files[0]

# Open the workbook
workbook = openpyxl.load_workbook(workbook_path)

# Print out all sheet names in the workbook for debugging
print("Available sheet names in the workbook:")
print(workbook.sheetnames)

# Function to write data to the workbook
def write_data_to_sheet(sheet, data, start_row, start_col, skip_header=False):
    for row_index, row in enumerate(data):
        if skip_header and row_index == 0:
            continue  # Skip the header row
        for col_index, value in enumerate(row):
            cell = sheet.cell(row=start_row + row_index, column=start_col + col_index)
            try:
                cell.value = float(value)
                cell.number_format = 'General'
            except ValueError:
                cell.value = value
            except Exception as e:
                print(f"Error writing value '{value}' to cell {cell.coordinate}: {e}")

# Get inputs for JAABA filenames to strip 
striping_JAABA_name1 = input("Enter the JAABA score file name to be removed: ")
striping_JAABA_name2 = input("Enter the JAABA score file name to be removed: ")

# Iterate through CSV files in the first folder
for csv_file in os.listdir(first_csv_folder):
    if csv_file.endswith(".csv"):
        csv_path = os.path.join(first_csv_folder, csv_file)
        csv_filename, _ = os.path.splitext(csv_file)
        sheet_name = csv_filename.replace("_output", "")  # Strip "_output" from CSV file name
        
        # Print derived sheet name for debugging
        print(f"Derived sheet name from first folder: {sheet_name}")
        
        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read data from CSV
            with open(csv_path, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)

            # Write data to the sheet starting from the second row
            write_data_to_sheet(sheet, data, start_row=2, start_col=1, skip_header=False)

        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")

# Iterate through CSV files in the second folder
for csv_file in os.listdir(second_csv_folder):
    if csv_file.endswith(".csv"):
        csv_path = os.path.join(second_csv_folder, csv_file)
        csv_filename, _ = os.path.splitext(csv_file)
        sheet_name = csv_filename.replace(striping_JAABA_name1, "").replace(striping_JAABA_name2, "")  # Strip any JAABA score file name from the final output name. 
        
        # Print derived sheet name for debugging
        print(f"Derived sheet name from second folder: {sheet_name}")
        
        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read data from CSV
            with open(csv_path, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)

            # Write data to the 3rd and 4th columns in the sheet, starting from the second row
            write_data_to_sheet(sheet, data, start_row=2, start_col=3, skip_header=False)

        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")

# Iterate through CSV files in the third folder
for csv_file in os.listdir(third_csv_folder):
    if csv_file.endswith(".csv"):
        csv_path = os.path.join(third_csv_folder, csv_file)
        csv_filename, _ = os.path.splitext(csv_file)
        sheet_name = csv_filename.replace("track_", "").split(".mp4")[0]  # Strip "track_" prefix and anything after ".mp4"
        
        # Print derived sheet name for debugging
        print(f"Derived sheet name from third folder: {sheet_name}")
        
        # Check if the sheet exists in the workbook
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Read data from CSV
            with open(csv_path, 'r') as f:
                reader = csv.reader(f)
                data = list(reader)

            # Write data to the 5th and 6th columns in the sheet, starting from the second row
            write_data_to_sheet(sheet, data, start_row=2, start_col=5, skip_header=False)

        else:
            print(f"Sheet '{sheet_name}' not found in the workbook.")

# Save the workbook
workbook.save(workbook_path)
