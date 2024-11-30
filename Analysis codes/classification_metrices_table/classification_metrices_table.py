# Last edited by Paulami Dey on 28/11/2024

import os
import openpyxl
import csv

# Function to find overlapping tuples
def finding_overlaptuples(test_list, test_tuples):
    result = []
    for test_tup in test_tuples:
        res = []
        for tup in test_list:
            if (tup[1] >= test_tup[0] and tup[0] <= test_tup[1]):
                res.append(tup)
        result.append(res)
    return result

# Replace 'your_file.xlsx' with the actual path to your Excel file
file_path = input("What is the file's path\n")

input_directory = os.path.dirname(file_path)

final_file_name = input("Enter the final file name for the output CSV file: ")
output_file = os.path.join(input_directory, f'output_{final_file_name}.csv')  # Output CSV file in the same directory as the script

try:
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    with open(output_file, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)

        print("Workbook loaded successfully.")

        # Collect sheet-wise metrics in a list
        all_metrics = []

        # Iterate over each sheet in the workbook
        for sheet_name in workbook.sheetnames:  
            sheet = workbook[sheet_name]

            print(f"Processing sheet: {sheet_name}")

            # Create empty lists to store data from each column
            A, B, C, D, E, F = [], [], [], [], [], []

            # Loop through each row of data (starting from row 2 to skip the header)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) >= 4:
                    A.append(row[0]) if row[0] is not None else None
                    B.append(row[1]) if row[1] is not None else None
                    C.append(row[2]) if row[2] is not None else None
                    D.append(row[3]) if row[3] is not None else None
                    E.append(row[4]) if row[4] is not None else None
                    F.append(row[5]) if row[5] is not None else None

            print("Data collected successfully.")

            # Zip the elements together
            Manual_list = list(zip(A, B))
            JAABA_list = list(zip(C, D))
            Matebook_list = list(zip(E, F))
            
            # Process the overlap counts for Manual vs JAABA
            false_positives = 0
            true_positives = 0
            false_negatives = 0
            result_data = []

           
            result_data.append([f"{sheet_name} - Counts for TP, FN: Manual Vs JAABA\n"])
            final_result = finding_overlaptuples(JAABA_list, Manual_list)
            for i, tup_list in enumerate(final_result):
                result_data.append(["Tuples overlapping with test tuple", Manual_list[i], "are:", tup_list])
                if not tup_list:
                    false_negatives += 1
                else:
                    true_positives += 1

            print("Processed Manual vs JAABA.")

            result_data.append([f"{sheet_name} - Counts for FP: Manual Vs JAABA\n"])
            final_result2 = finding_overlaptuples(Manual_list, JAABA_list)
            for i, tup_list in enumerate(final_result2):
                result_data.append(["Tuples overlapping with test tuple", JAABA_list[i], "are:", tup_list])
                if not tup_list:
                    false_positives += 1

            print("Processed Manual vs JAABA.")

            false_positives_MB = 0
            true_positives_MB = 0
            false_negatives_MB = 0

            result_data.append([f"{sheet_name} - Counts for TP, FN: Manual Vs Matebook\n"])
            final_result3 = finding_overlaptuples(Matebook_list, Manual_list)
            # Printing result
            for i, tup_list in enumerate(final_result3):
                result_data.append(["Tuples overlapping with test tuple", Manual_list[i], "are:", tup_list])
                if not tup_list:
                    false_negatives_MB += 1
                else:
                    true_positives_MB += 1

            print("Processed Manual vs Matebook.")

            result_data.append([f"{sheet_name} - Counts for FP: Manual Vs Matebook\n"])
            final_result4 = finding_overlaptuples(Manual_list, Matebook_list)
            # Printing result
            for i, tup_list in enumerate(final_result4):
                result_data.append(["Tuples overlapping with test tuple", Matebook_list[i], "are:", tup_list])
                if not tup_list:
                    false_positives_MB += 1  

            print("Processed Manual vs Matebook.")

            # Calculate precision, recall, and F1 score
            precision = 0 if (true_positives + false_positives) == 0 else (true_positives) / (true_positives + false_positives)
            recall = 0 if (true_positives + false_negatives) == 0 else (true_positives) / (true_positives + false_negatives)
            F1_score = 0 if (precision + recall) == 0 else 2 * (precision * recall) / (precision + recall)

            precision_MB = 0 if (true_positives_MB + false_positives_MB) == 0 else (true_positives_MB)/(true_positives_MB + false_positives_MB)
            recall_MB = 0 if (true_positives_MB + false_negatives_MB) == 0 else (true_positives_MB)/(true_positives_MB + false_negatives_MB)
            F1_score_MB = 0 if (precision_MB + recall_MB) == 0 else 2*(precision_MB*recall_MB)/(precision_MB + recall_MB)

            # Append metrics for this sheet to all_metrics list
            all_metrics.append([sheet_name, true_positives ,false_positives, false_negatives, precision, recall, F1_score, true_positives_MB ,false_positives_MB, false_negatives_MB, precision_MB, recall_MB, F1_score_MB])

            print("Metrics calculated.")

        # Write all_metrics to CSV file
        writer.writerow(["Sheet Name", "True postives", "False postives", "False negatives", "Precision", "Recall", "F1 Score", "True positives (MB)", "False positives (MB)", "False negatives (MB)", "Precision (Matebook)", "Recall (Matebook)", "F1 Score (Matebook)"])
        writer.writerows(all_metrics)

        print("Data written to CSV successfully.")

except FileNotFoundError:
    print("File not found. Please check the file path.")
except Exception as e:
    print("An error occurred:", e)
